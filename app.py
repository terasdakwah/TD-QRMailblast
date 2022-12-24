import os, json, requests, pika, time, datetime
from config import RABBITMQ_HOST, RABBITMQ_USER, RABBITMQ_PASS
from flask import Flask, render_template
from flask import flash, request, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app = Flask(__name__)
app.secret_key = "secret key"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1000 * 1000

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def send_email_queue(name, email, judul):
    queue = 'email-qr'
    connection = pika.BlockingConnection(pika.ConnectionParameters(
        host=RABBITMQ_HOST,
        port=5672,
        virtual_host='/',
        credentials=pika.credentials.PlainCredentials(
            username=RABBITMQ_USER,
            password=RABBITMQ_PASS
        )
    ))

    channel = connection.channel()
    channel.queue_declare(queue=queue, durable=True)

    timestamp = time.time()
    now = datetime.datetime.now()
    expire = 10000 * int((now.replace(hour=23, minute=59, second=59, microsecond=999999) - now).total_seconds())
    headers = {
        'code': '200',
        'created': int(timestamp)
    }
    data = {
        'name': str(name),
        'email': str(email),
        'judul': str(judul),
        'created': int(timestamp),
        'expire': expire
    }
    channel.basic_publish(
        exchange='',
        routing_key=queue,
        body=json.dumps(data),
        properties=pika.BasicProperties(
            delivery_mode=2,
            priority=0,
            timestamp=int(timestamp),
            expiration=str(expire),
            headers=headers
        ))
    connection.close()
    return True

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    req_rabbitmq = requests.get('http://%s:15672/api/queues' % (RABBITMQ_HOST,), auth=(RABBITMQ_USER, RABBITMQ_PASS))
    result_rabbitmq = json.loads(req_rabbitmq.text)
    messages_ready = result_rabbitmq[0].get('messages_ready',0)

    if messages_ready > 0:
        return render_template('waiting.html', sisa=messages_ready)

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('File tidak ditemukan')

        file = request.files['file']
        if file.filename == '':
            flash('File belum dipilih')

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            fileurl = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(fileurl)
            
            try:
                wb = load_workbook(filename = fileurl)
                sheet_ranges = wb['email']
                index = 0

                if '@' in sheet_ranges['C2'].value:
                    for row in sheet_ranges.iter_rows():
                        index=index+1
                        if index==1 or row[2].value==None or '@' not in row[2].value:
                            continue

                        judul = request.form.get('judul', '')
                        if judul=='':
                            judul = 'QR Kajian Terasdakwah'
                        send_email_queue(row[1].value, row[2].value, judul)

                    flash('Proses mengirim email')
                    flash('Refresh halaman ini untuk melihat prosesnya')
                    return redirect(url_for('upload_file'))
                else:
                    flash('Format file salah')

            except Exception as e:
                print(e)
                flash('Format file salah')

        else:
            flash('Format file salah wajib .xlsx')

    return render_template('app.html')